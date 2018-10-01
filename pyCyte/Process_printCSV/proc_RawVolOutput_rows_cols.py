# -*- coding: utf-8 -*-
"""
Created on Sat Apr  7 19:35:29 2018

@author: ljungherr
"""
import os
import matplotlib.pyplot as plt
import numpy as numpy
import pyCyte.ToolBox.SimpleTools as sbox
#readers  = [ 'Biotek1', 'Biotek2', 'BMG_1310', 'BMG_1619' ]
#curves = ['ul', 'ul_DMSO', 'ul_GP']
types = ['Flat','Other', 'FineES', 'RRSweep' ]
patterns = ['rows', 'columms', 'UCP'] ; pattern = 'UCP'
lots = ['Lot1', 'Lot2', 'Lot3'] ; 
xLimits = [[]]
colors = ['green', 'blue', 'purple', 'red', 'brown', 'pink',  'orange','lightblue','yellow','cyan',
     'mediumblue','lightgreen','forestgreen', 'yellow','orange','aliceblue','lavender','violet']
Scatter = True ; EjectSweep = True ; FillThick = False ; Single = False; cIndex = -1

subdir = '20180417_96PPQ_Lot_Prints\ReaderFiles_ES'
os.chdir('''\\\\seg\\data\\96SP_Feasibility\\Experiments\\{0}\\'''.format(subdir))
#os.chdir ('\\\\seg\\data\\96SP_Feasibility\\Experiments\\20180411_96PPQ_SysCal_Lot1\96PPQ_AQ_BP\EjectSweep')

srcRows = []
for letter in range (65,73):
    srcRows.append(chr(letter))
    
pwr = numpy.arange(-0.44, 0.52,0.04)
fillVols = [60,100,140,180,200,220,260,300]
from openpyxl import load_workbook
fl = sbox.getFileList('*RawVolumeOutput.xlsx')
        
plt.figure(figsize = (12,7.5))
plt.suptitle ('96PPQ - 3 Resin Lots - Print Volume vs. Fill Volume', fontsize = 15)

#for sub in range (1,5):
#    plt.subplot(2,2,sub)
#    plt.title('''Type = {0}'''.format (types[sub-1]))
#    plt.xlabel (xlabels[sub - 1], fontsize = 14); plt.ylim(90,120)
#    plt.ylim(-.15, 0.05); plt.xlim(0,9900)

if EjectSweep:
    xLabel = 'Power Adjust (db)'  ;  xLim = [-0.5, 0.5] ; yLim = [75,115]
if FillThick:
    xLabel = 'Src Volume (uL)'     ; xLim = [0, 350] ; yLim = [85,110]
for f in fl:
    fsplit = f.split ('\\'); 
    for i in range (0,len(fsplit)):
      if '2018' in fsplit[i]:
        time = fsplit[i][9:15]
        ftime = formatTime(time)
    Tag = plt.scatter ([],[], label = ftime, color = colors[cIndex])

    wb2 = load_workbook(f, data_only = True)
    sheets = wb2.get_sheet_names()

#        print (thisLot, plateName)
                   
#      thisLegend = '''{1}  {2} '''.format(lot, plateName )
#      sheet = wb2[sh]
#      xValues = [] ;  yValues = [] ; errors = []; stDevs = []
      
#      for col in range (2,26):
#            xv = sheet.cell(row = 20, column = col).value
#            yv = sheet.cell(row = 21, column = col).value
#            errpct = sheet.cell(row = 25, column = col).value
#            std = sheet.cell(row = 22, column = col).value
#            xValues.append (xv) ; yValues.append(yv)
#            errors.append(errpct) ; stDevs.append (std)
##      print (readers.index(thisR), thisR, thisLegend , nom, obs)
#      plt.subplot(2,2, types.index(type) + 1 )
#      Tag = plt.plot ([],[], label = thisLegend, color = colors[sheets.index(sh)], lw = 2)
#      plt.plot (xValues, yValues, color = colors[sheets.index(sh)])

#for sub in range (1,5):
#    plt.subplot(2,2,sub)
#    plt.legend (loc = 2, fontsize = 11 ) 
    for sh in sheets:
      if not 'Lot2' in sh:
        cIndex += 1
        for lot in lots:
            if lot in sh:
                thisLot = lot
        plateName = sh[0:7]
        
#        plt.subplot(2,3, sheets.index(sh) +1)
        plt.subplot (2,2, cIndex+1)
        plt.title (plateName, fontsize = 14)
        if cIndex > 1:
            plt.xlabel (xLabel)
        if cIndex == 0 or cIndex == 2:
            plt.ylabel ('Print Volume (nL)')
        plt.ylim(yLim); plt.xlim (xLim)

        thisSheet = wb2[sh]
        allRows =[] ;  AvgVols = [] ; StDevs = []
        for row in thisSheet.iter_rows('A2:Y17'):
            destRow = row[0].value
            rowList = []
            for col in range(1,25):
                cellVal = float(row[col].value)
                rowList.append (cellVal)
            allRows.append (rowList)  #allRows is used to plot data vs column variable.
			
            rowMean = numpy.mean(rowList)
            rowStd = numpy.std(rowList)
            rowDict[destRow] = rowList ; rowDict['rMean'] = rowMean; rowDict['rStd'] = rowStd

        for sr in srcRow:
#        print (sr, colors[srcRow.index(sr)])
            destRow1 = []; destRow2 = []; destRowBoth = []
            destRow1 = allRows[srcRow.index(sr) * 2]
            destRow2 = allRows[srcRow.index(sr) * 2 + 1]
            destRowBoth.extend(destRow1)
            destRowBoth.extend(destRow2)

            if EjectSweep:   
                plt.scatter (pwr, destRow1[0:24], color = colors[srcRow.index(sr)])
                plt.scatter (pwr, destRow2, color = colors[srcRow.index(sr)])
            thisMean = numpy.mean(destRowBoth)
            thisStd = numpy.std(destRowBoth)
            AvgVols.append (thisMean) ; StDevs.append (thisStd)
        if FillThick:
           print(''' Sheet = {0}  Source Row = {1}'''.format (plateName, sr))
#
           plt.scatter (fillVols, AvgVols , color = colors[srcRow.index(sr)])
           plt.errorbar(fillVols, AvgVols,yerr = StDevs, linestyle="None")
        
        
def formatTime(time):
    if not len(time) == 6 and int(time):
        print ('Expecting a 6 digit-string for time. Returning string with no changes.')
        return (time)
        
    hour = time[0:2]
    mins = time[2:4]
    sec = time[4:6]
    formattedTime = '''{0}-{1}-{2}'''.format(hour, mins, sec)
    return(formattedTime)
        