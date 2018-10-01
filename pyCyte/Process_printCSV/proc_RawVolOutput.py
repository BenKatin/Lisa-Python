# -*- coding: utf-8 -*-
"""
Created on Sat Apr  7 19:35:29 2018

@author: ljungherr
"""
import os
import matplotlib.pyplot as plt
import numpy as numpy
import pyCyte.ToolBox.SimpleTools as sbox
#readers  = [ 'Biotek1', 'Biotek2', 'BMG_1310', 'BMG_1619' ]
#curves = ['ul', 'ul_DMSO', 'ul_GP']
types = ['Flat','Other', 'FineES', 'RRSweep' ]
xlabels = ['','', 'Power (db)', 'RR (Hz)', ]
xLimits = [[]]
colors = ['red','orange', 'green', 'lightgreen','blue', 'lightblue','purple', 'brown', 'black','pink','cyan','lightgreen','forestgreen', 'yellow','orange','aliceblue','lavender','violet']
os.chdir ('\\\\seg\\data\\96SP_Feasibility\\Experiments\\20181410_96PPRE_Stability')

from openpyxl import load_workbook
fl = sbox.getFileList('*RawVolumeOutput.xlsx')
        
plt.figure(figsize = (14,8))

for sub in range (1,5):
    plt.subplot(2,2,sub)
    plt.title('''Type = {0}'''.format (types[sub-1]))
    plt.xlabel (xlabels[sub - 1], fontsize = 14); plt.ylim(90,120)
#    plt.ylim(-.15, 0.05); plt.xlim(0,9900)
        
for f in fl:

    wb2 = load_workbook(f, data_only = True)
    sheets = wb2.get_sheet_names()
    for sh in sheets:
      OffsetType = 'Xoff'
      if 'Yoff' in sh:
          OffsetType = 'Yoff'
      if '_31_' in sh:
           plateName = 'Plate 31'; p = 0
      if '_42_' in sh:
           plateName = 'Plate 42'; p = 1
      if '_18_' in sh:
          plateName = 'Plate 18'; p = 2
      if '_32_' in sh:
          plateName = 'Plate 32'; p = 3
      MIPType = 'MIP Off'
      if 'MIP' in sh:
          MIPType = 'MIP On'
      type = 'Other'
      for t in types:
          if t in sh:
              type = t
    
#    book = openpyxl.load_workbook('sample.xlsx')
      thisLegend = '''{1}  {2}  {3}'''.format(type, plateName, OffsetType, MIPType)
      sheet = wb2[sh]
      xValues = [] ;  yValues = [] ; errors = []; stDevs = []
      for col in range (2,26):
            xv = sheet.cell(row = 20, column = col).value
            yv = sheet.cell(row = 21, column = col).value
            errpct = sheet.cell(row = 25, column = col).value
            std = sheet.cell(row = 22, column = col).value
            xValues.append (xv) ; yValues.append(yv)
            errors.append(errpct) ; stDevs.append (std)
#      print (readers.index(thisR), thisR, thisLegend , nom, obs)
      plt.subplot(2,2, types.index(type) + 1 )
      Tag = plt.plot ([],[], label = thisLegend, color = colors[sheets.index(sh)], lw = 2)
      plt.plot (xValues, yValues, color = colors[sheets.index(sh)])

for sub in range (1,5):
    plt.subplot(2,2,sub)
    plt.legend (loc = 2, fontsize = 11 ) 

    