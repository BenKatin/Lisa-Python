 # -*- coding: utf-8 -*-
"""
Created on Thu Aug 24 11:10:46 2017

@author: avandenbroucke
       : lchang - Added EPCFT functionality in V4

see also https://stackoverflow.com/questions/34276663/tkinter-gui-layout-using-frames-and-grR = R.sortlevel(['Fluid','Plastic','Concentration'], sort_remaining=False )id
"""

print(' Starting Application ' )

import time
import datetime
print('imported time/datetime @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )

import sys
if not('.' in sys.path): sys.path.append('.')
#import midi24txt 
print('imported sys @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )

import tkinter 
from tkinter import ttk
#from tkFileDialog import *
from tkinter import filedialog
from tkinter import messagebox
print('imported tkinter @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )

import pandas
#import numpy as np
#import scipy.stats
print('imported pandas @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )

import matplotlib.pyplot as plt
print('imported pyplot @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )
import os
print('imported os @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )
if __package__ is None or __package__ == '':
    from pyCyte.DropCV import ReaderDat
    from pyCyte.DropCV import processCCT
    from pyCyte.EjectSweep import processEjectSweep
    from pyCyte.DropCV import ProcessEPCFinalTest as EPCFT
else:
    from ..DropCV import ReaderDat
    from ..DropCV import processCCT
    from ..EjectSweep import processEjectSweep
    from ..DropCV import ProcessEPCFinalTest as EPCFT
print('imported pyCyte @ ', datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S.%f') )
#from PIL import ImageTk, Image
# thinking in tkinter http://www.ferg.org/thinking_in_tkinter/all_programs.html
import subprocess
import openpyxl
from openpyxl import formatting, styles
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


class TheGui:
    def __init__(self, parent):
        self.ReaderDat = ReaderDat.ReaderDat()
        self.verbose = True
        self.ReaderDat.verbose = self.verbose
        self.wdir = os.getcwd()
        self.tdir = os.getcwd()
        self.processES = processEjectSweep.processEjectSweep()
        self.sweepData = True
        
        self.readMode = tkinter.BooleanVar()
        self.readMode.set(True)
        self.readModeColor = tkinter.StringVar()
        self.readModeColor.set('chartreuse')
        self.curveMode = tkinter.BooleanVar()
        self.curveMode.set(True)
        self.curveModeColor = tkinter.StringVar()
        self.curveModeColor.set('chartreuse')        
        
        #==== TransferData ======
        self.transferData = False
        self.tableIndex = tkinter.IntVar()
        self.tableIndex.set(0)
        self.FTfolder = tkinter.StringVar()
        self.loadedtransferfiles = []
       
        # allows to size the window
        ttk.Sizegrip(root).grid(column=999, row=999, sticky=('SE'))

        # define some widths
        self.overallwidth = 860        
        self.fwidth = 450
        self.top_frameheight=150 
        self.file_frameheight = 250
        self.listframeheight = self.file_frameheight - 5
        self.listframewidth = 650
        self.btm_frameheight = 145
        self.center_frameheight = 150     
        
        # create all of the main containers
        self.top_frame = tkinter.Frame(parent,  width=self.overallwidth, height=200, pady=3)
        self.file_frame = tkinter.Frame(parent, borderwidth=5, relief='ridge', width=self.overallwidth,height=200,padx=3 )
        self.center_frame = tkinter.Frame(parent,width=self.overallwidth, height=self.center_frameheight, padx=3)
        self.btm_frame = tkinter.Frame(parent, width=self.overallwidth, height=self.btm_frameheight, pady=3)
        self.top_settings_frame = tkinter.Frame(parent, borderwidth=5, relief='sunken', width=300, height=100, pady=5)
        self.center_FT_frame = tkinter.Frame(parent, borderwidth=5, relief='ridge', width=self.overallwidth-250, height=self.center_frameheight, pady=10)
        
        # layout main containers
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        
        # put main containers on grid
        self.top_frame.grid(row=0, sticky="nsew")
        self.file_frame.grid(row=1,sticky="nsew")
        self.center_frame.grid(row=2, sticky="nsew")
        self.btm_frame.grid(row=3, pady=10, sticky="nsew")
        self.top_settings_frame.grid(row=0, sticky = "ne", padx = 5, pady=5)
        self.center_FT_frame.grid(row=2, sticky = "ne")
  
        #Scrollbar configurate, attach the scrollbar to "listFrame", 
        #when the size of frame "ListFrame" exceed the size of ListFrame, the scrollbar acts
        #( from https://stackoverflow.com/questions/19245941/adding-a-scrollbar-to-a-frame-using-tkinter-python )
        self.canvas=tkinter.Canvas(self.file_frame, borderwidth=0)
        self.listFrame=tkinter.Frame(self.canvas)#, bg='green')
        self.scrollb=tkinter.Scrollbar(parent, orient="vertical",command=self.canvas.yview)
        self.scrollb.grid(row=1, column=1, sticky='nsew')  #grid scrollbar in master, but
        self.canvas['yscrollcommand'] = self.scrollb.set   #attach scrollbar to FILE_fRAME
    #    self.scrollb.pack(side="right",fill="y")      
  
   #     
        self.canvas.create_window((4,4),window=self.listFrame,anchor='nw')  
        self.listFrame.bind("<Configure>", self.AuxscrollFunction)
     #   self.file_frame.bind("<Configure>", lambda event, scanvas=self.canvas: self.onFrameConfigure(self.canvas))
        self.scrollb.grid_forget()    
        self.canvas.pack(side="left",fill="both",expand=True)
        
        
        #------- Top Frame  ----------#
        # variables assigned in top frame
        self.useInternalStd = tkinter.BooleanVar()
        self.useInternalStd.set('False')
        
        self.applyPMTCorr = tkinter.BooleanVar()
        self.applyPMTCorr.set('True')
        
        self.exclude16 = tkinter.BooleanVar()
        self.exclude16.set('False')
        
        self.rawVolOutput = tkinter.BooleanVar()
        self.rawVolOutput.set('True')
        
        self.supportedReaders = self.getSupportedReaders()
        self.selectedReader = tkinter.StringVar()
        self.selectedReader.set( 'BioTek2' )
        
        self.supportedLUTs = self.getSupportedLUTs()
        print(self.supportedLUTs)
        self.selectedLUT = tkinter.StringVar()
        self.selectedLUT.set( self.supportedLUTs[0])
        
        self.INSTRUMENTS = [ 'E6XX/E55X', 'E525']
        self.echoType = tkinter.StringVar()
        self.echoType.set(self.INSTRUMENTS[0])
        
        self.CRITERIA =  [ ('MFG Specs','MFG'),('Customer Specs','CUST'),('Service Specs','SER')]
        self.passfailCriteria = tkinter.StringVar()
        self.passfailCriteria.set(self.CRITERIA[0][1])                  
        
        self.MODES = [ ("FinalTest","FT"), ("CCT","CCT") ]
        self.operationMode = tkinter.StringVar()
        self.operationMode.set(self.MODES[0][1])
        
        self.RUNTYPES = ['Print','Sweep','EPCFT']
        self.runType = tkinter.StringVar()
        self.runType.set(self.RUNTYPES[0])
               
        self.ReaderFileList = []
        
        #widgets top frame
        #Sweep vs Print
        i = 0
        for rmode in self.RUNTYPES:
            self.t1 = tkinter.Radiobutton(self.top_frame, text=rmode, width='10', variable = self.runType, value = rmode, command=self.selrtype, indicatoron = 0, bg = 'light sky blue')
            self.t1.grid(row=0,column=i,columnspan=2,padx="5", sticky="w")
            i +=2        
         # Tests per Sweep 
        self.nTestsLab = tkinter.Label(self.top_frame, width='16', text = '# SweepTestsPerPlate: ')
        self.numSplitsPerSweep = tkinter.Spinbox(self.top_frame,width='4',from_=0, to_=99 )
        self.numSplitsPerSweep.delete(0,'end')
        self.numSplitsPerSweep.insert(0,1) 
        
        # Load FT folder
        self.btnLoadFT = tkinter.Button(self.top_frame, width = '10', text = 'FT Folder',command=self.loadFTFolder)
        
        #InternalStandardCheckBox
        self.c1 = tkinter.Checkbutton(self.top_frame, width='10', text="Use IntStd", variable=self.useInternalStd,  command=self.selc1)
        #Apply PMT correction CheckBox
        self.c2 = tkinter.Checkbutton(self.top_frame, width='10', text="Use PMTCorr", variable=self.applyPMTCorr)
        self.c3 = tkinter.Checkbutton(self.top_frame, width='10', text="Ex.Col16", variable=self.exclude16, command=self.selc3)
        self.c4 = tkinter.Checkbutton(self.top_frame, width='10', text='RawVolumeFile',variable=self.rawVolOutput)

        #Reader label and OptionMenu
        self.lm1 = tkinter.Label(self.top_frame, width='5', text='Reader:')        
        self.m1 = tkinter.OptionMenu(self.top_frame, self.selectedReader, *self.supportedReaders,  command=self.setLUT)
        self.m1.config(width='25')
        #LUT label and OptionMenu
        self.lm2 = tkinter.Label(self.top_frame, width='5',text='LUT:')
        self.m2 = tkinter.OptionMenu(self.top_frame,  self.selectedLUT, *self.supportedLUTs,  command=self.printLUT)
        
        self.m1.grid(row=0,column=i+2,padx=5,sticky="ew")
        self.lm1.grid(row=0,column=i+1)
        self.m2.grid(row=1,column=i+2,padx=5,sticky="ew")
        self.lm2.grid(row=1,column=i+1)
        
         #Reader File list
#        self.lblInRF = tkinter.Label(self.center_frame, text='Reader Files:', anchor='w',width='15')
        self.btnInLF = tkinter.Button(self.top_frame,width='25', bg='white', text='Load Reader Files', command=self.btnSelectFiles)
 #       self.frmfilebox = tkinter.Listbox(self.center_frame, width=20, height=1)
 #       self.frmfilebox.insert(tkinter.END, *self.ReaderFileList)
        # Echo SN
        self.snlab = tkinter.Label(self.top_frame, width='8', text = 'EchoSN: ')
        self.EchoSN = tkinter.Entry(self.top_frame,width='8' )
        self.EchoSN.insert(0,'E17YY')
 
        self.btnUpdateAttr = tkinter.Button(self.top_frame,width='10', bg='white', text='Update Attributes', command=self.goUpdateAttributes)
        self.btnUpdateAttr.config(state='disabled')
        
        # modes radiobutton:
        j=0     
        for text, mode in self.MODES:
            self.b1 = tkinter.Radiobutton(self.top_frame, text=text, width='8', variable=self.operationMode, value=mode , indicatoron = 0, command=self.btnOperationMode )
            self.b1.grid(row=j, column=i+3, padx="5", sticky="w")   
            j += 1
        # instruments radiobutton:
        k=0     
        for instrument in self.INSTRUMENTS:
            self.b2 = tkinter.Radiobutton(self.top_frame, text=instrument, width='8', variable=self.echoType, value=instrument , indicatoron = 0 )
            self.b2.grid(row=k, column=i+4, padx="5", sticky="w")   
            k += 1
         #passfail criteria
        m=0     
        for text, mode  in self.CRITERIA:
            self.b3 = tkinter.Radiobutton(self.top_frame, text=text, width='12', variable=self.passfailCriteria, value=mode , indicatoron = 0 )
            self.b3.grid(row=m, column=i+7, padx="5", sticky="w")   
            m += 1
            
        self.snlab.grid(row=0,column=i+5,padx=5,sticky="ew")
        self.EchoSN.grid(row=0,column=i+6,padx=5,sticky="ew")
        self.nTestsLab.grid(row=1,column=0,columnspan=3,padx=10,sticky="e")
        self.numSplitsPerSweep.grid(row=1,column=3,padx=10,sticky="e")
        self.btnLoadFT.grid(row=1, column=4, padx=10, sticky="e")
        self.c1.grid(row=2,column=0,columnspan=2,sticky="ew")
        self.c2.grid(row=2,column=2,columnspan=2,sticky="ew")
        self.c3.grid(row=3,column=0,columnspan=2,sticky="w")
        self.c4.grid(row=3,column=2,columnspan=3,sticky="w")
        

        # Load File button
  #      self.lblInRF.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
        self.btnInLF.grid(row=2,column=i+2,columnspan=5, padx=5,sticky="ew")
        self.btnUpdateAttr.grid(row=1, column=i+5,columnspan=2,padx=5,sticky="ew")
        
        #------ Settings Frame --------#
        self.nomVol=tkinter.StringVar()
        self.nomVol.set(50.0)
        
        self.readModeLabelStart = tkinter.Label(self.top_settings_frame, text = 'Fluor', padx=5, pady=5)
        self.readModeLabelEnd = tkinter.Label(self.top_settings_frame, text = 'Abs', padx=5, pady=5)
        self.curveModeLabelStart = tkinter.Label(self.top_settings_frame, text = 'MFG', padx=5, pady=5)
        self.curveModeLabelEnd = tkinter.Label(self.top_settings_frame, text = 'R&D', padx=5, pady=5)
        self.readModeScale = tkinter.Scale(self.top_settings_frame, command = self.updateReadMode, from_=0, to=1, orient='horizontal', showvalue=False, length=50, troughcolor=self.readModeColor.get())
        self.curveModeScale = tkinter.Scale(self.top_settings_frame, command = self.updateCurveMode,from_=0, to=1, orient='horizontal', showvalue=False, length=50, troughcolor=self.curveModeColor.get())
        self.nomVolLabelStart = tkinter.Label(self.top_settings_frame, text = 'NomVol:', padx=5, pady=5)
        self.nomVolLabelEnd = tkinter.Label(self.top_settings_frame, text = 'nL', padx=5, pady=5)
        self.nomVolEnt = tkinter.Spinbox(self.top_settings_frame,width=6, from_=0.0, to_=10000.0, textvariable = self.nomVol, format = '%.1f', increment = 0.1)
        self.outlierLabelStart = tkinter.Label(self.top_settings_frame, text = 'Outlier:', padx=5, pady=5)
        self.outlierLabelEnd = tkinter.Label(self.top_settings_frame, text = '% avg vol', padx=5, pady=5)
        self.outlierEnt = tkinter.Spinbox(self.top_settings_frame, width=6, from_=0, to_=100)
        self.outlierEnt.delete(0,'end')
        self.outlierEnt.insert(0,15)
        self.relLabelStart = tkinter.Label(self.top_settings_frame, text = 'Reliability:', padx=5, pady=5)
        self.relLabelEnd = tkinter.Label(self.top_settings_frame, text = '% avg vol', padx=5, pady=5)
        self.relEnt = tkinter.Spinbox(self.top_settings_frame, width=6, from_=0, to_=100)
        self.relEnt.delete(0,'end')
        self.relEnt.insert(0,40)
        
        self.nomVolLabelStart.grid(row=0, column=0, sticky='w')
        self.nomVolLabelEnd.grid(row=0, column=2, sticky='w')
        self.nomVolEnt.grid(row=0, column=1, sticky='w')
        self.outlierLabelStart.grid(row=1, column=0, sticky='w')
        self.outlierEnt.grid(row=1, column=1, sticky='w')
        self.outlierLabelEnd.grid(row=1, column=2, sticky='w')
        self.relLabelStart.grid(row=2, column=0, sticky='w')
        self.relEnt.grid(row=2, column=1, sticky='w')
        self.relLabelEnd.grid(row=2, column=2, sticky='w')
        self.readModeLabelStart.grid(row=1, column=3, sticky='w')
        self.readModeLabelEnd.grid(row=1, column=5, sticky='w')
        self.readModeScale.grid(row=1, column=4, sticky='w')
        self.curveModeLabelStart.grid(row=0, column=3, sticky='w')
        self.curveModeLabelEnd.grid(row=0, column=5, sticky='w')
        self.curveModeScale.grid(row=0, column=4, sticky='w')
        
        #Disabled for now
        self.outlierEnt.configure(stat='disabled')
        self.relEnt.configure(stat='disabled')
        self.outlierLabelStart.configure(stat='disabled')
        self.outlierLabelEnd.configure(stat='disabled')
        self.relLabelStart.configure(stat='disabled')
        self.relLabelEnd.configure(stat='disabled')
        self.readModeLabelStart.config(state='disabled')
        self.readModeLabelEnd.config(state='disabled')
        self.readModeScale.config(state='disabled')
        
        #------- Center Frame  ----------#
        #variables assigned in Center Frame
        self.ReaderFileAttributes = {}
        self.MatechFileName1 = tkinter.StringVar() # http://effbot.org/tkinterbook/entry.htm
        self.MatechFileName2 = tkinter.StringVar()
        self.PickListFileName = tkinter.StringVar()
        self.transferFileName = tkinter.StringVar()
        self.ReaderFileList = []
        self.OutPutFolder = tkinter.StringVar() # http://effbot.org/tkinterbook/entry.htm
        self.OutPutFolder.set(self.wdir)
        self.displayTransferFiles = tkinter.StringVar()
        
        #------- FT Frame  ----------#
        #variables assigned in FT Frame
        self.surveyFileName = tkinter.StringVar()
        self.printFileName = tkinter.StringVar()
        self.MIPOutFileName = tkinter.StringVar()
#        self.MIPInfoFileName = tkinter.StringVar()
                
        #widget center frame
        lblwidth = 25
        entwidth = 100
        btnwidth = 12
        #Matech File 1
        self.lblIn1 = tkinter.Label(self.center_frame, text='Matech File 1:', anchor='w',width=lblwidth)      
        self.entIn1 = tkinter.Entry(self.center_frame, width=entwidth, textvariable=self.MatechFileName1)
        self.btnIn1 = tkinter.Button(self.center_frame, width=btnwidth, text='Browse', command=self.btnInBrowseClick1)
        #Matech File 2
        self.lblIn2 = tkinter.Label(self.center_frame, text='Matech File 2:', anchor='w',width=lblwidth)
        self.entIn2 = tkinter.Entry(self.center_frame,  width=entwidth, textvariable=self.MatechFileName2)
        self.btnIn2 = tkinter.Button(self.center_frame,  width=btnwidth, text='Browse', command=self.btnInBrowseClick2)
        #PickList File
        self.lblIn3 = tkinter.Label(self.center_frame, text='PickList File (Optional)', anchor='w', width=lblwidth)
        self.entIn3 = tkinter.Entry(self.center_frame, width=entwidth, textvariable=self.PickListFileName)
        self.btnIn3 = tkinter.Button(self.center_frame, width=btnwidth, text='Browse', command=self.btnSelectPickListFile)
        #Output Folder
        self.lblOut1 = tkinter.Label(self.center_frame, text='Output folder:', anchor='w',width=lblwidth)       
        self.entOut1 = tkinter.Entry(self.center_frame, width=entwidth, textvariable=self.OutPutFolder)
        self.btnOut1 = tkinter.Button(self.center_frame, width=btnwidth,text='Browse', command=self.btnSelectOutputFolder1) 
       
        #FinalTest Information
        self.lblFT1 = tkinter.Label(self.center_FT_frame, text = 'Final Test Survey Files:', anchor = 'w', width = lblwidth)
        self.entFT1 = tkinter.Entry(self.center_FT_frame, width=entwidth-10, textvariable=self.surveyFileName)
        self.btnFT1 = tkinter.Button(self.center_FT_frame, width=btnwidth, text = 'Browse', command =self.btnSurveyClicked)
        
        self.lblFT2 = tkinter.Label(self.center_FT_frame, text = 'Final Test Print Files:', anchor = 'w', width = lblwidth)
        self.entFT2 = tkinter.Entry(self.center_FT_frame, width=entwidth-10, textvariable=self.printFileName)
        self.btnFT2 = tkinter.Button(self.center_FT_frame, width=btnwidth, text = 'Browse', command =self.btnPrintClicked)
        
        self.lblFT3 = tkinter.Label(self.center_FT_frame, text = 'Final Test MIPOutput Files:', anchor = 'w', width = lblwidth)
        self.entFT3 = tkinter.Entry(self.center_FT_frame, width=entwidth-10, textvariable=self.MIPOutFileName)
        self.btnFT3 = tkinter.Button(self.center_FT_frame, width=btnwidth, text = 'Browse', command =self.btnMIPOutClicked)

#        self.lblFT4 = tkinter.Label(self.center_FT_frame, text = 'Final Test MIPInfoEX Files:', anchor = 'w', width = lblwidth)
#        self.entFT4 = tkinter.Entry(self.center_FT_frame, width=entwidth-10, textvariable=self.MIPInfoFileName)
#        self.btnFT4 = tkinter.Button(self.center_FT_frame, width=btnwidth, text = 'Browse', command =self.btnMIPInfoClicked)        
        
        rowcnt = 0
        span = 1
        self.lblIn1.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
        self.btnIn1.grid(row=rowcnt,column=1,padx=5,sticky="e")
        self.lblFT1.grid(row=rowcnt,column=0,padx=40,pady=10,sticky="w")
        self.btnFT1.grid(row=rowcnt,column=1,padx=40,sticky="e")

        
        rowcnt +=span
        self.entIn1.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=10,pady=10,sticky="nsew")
        self.entFT1.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=40,pady=10,sticky="nsew")
        
        rowcnt += span
        self.lblIn2.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
        self.btnIn2.grid(row=rowcnt,column=1, padx=5, sticky="e")
        self.lblFT2.grid(row=rowcnt,column=0,padx=40,pady=5,sticky="w")
        self.btnFT2.grid(row=rowcnt,column=1, padx=40, sticky="e")
    
        rowcnt +=span 
        self.entIn2.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=5,pady=5,sticky="nsew")
        self.entFT2.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=40,pady=10,sticky="nsew")
        
        rowcnt += span
        self.lblIn3.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
        self.btnIn3.grid(row=rowcnt,column=1, padx=5, sticky="e")
        self.lblFT3.grid(row=rowcnt,column=0,padx=40,pady=10,sticky="w")
        self.btnFT3.grid(row=rowcnt,column=1, padx=40, sticky="e")
        
        rowcnt += span
        self.entIn3.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=5,pady=5,sticky="nsew")
        self.entFT3.grid(row=rowcnt,column=0,columnspan=2,rowspan=span,padx=40,pady=10,sticky="nsew")
        
        rowcnt += span 
        self.lblOut1.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
        self.btnOut1.grid(row=rowcnt,column=1, padx=5,sticky="e")
#        self.lblFT4.grid(row=rowcnt,column=0,padx=5,pady=5,sticky="w")
#        self.btnFT4.grid(row=rowcnt,column=1, padx=5,sticky="e")

        rowcnt +=span
        self.entOut1.grid(row=rowcnt,column=0,columnspan=2,padx=5,pady=5,sticky="ew")
#        self.entFT4.grid(row=rowcnt,column=0,columnspan=2,padx=5,pady=5,sticky="ew")

       
        
  
        #-------- File Frame ---------#
        #variables
        self.optionmenu_plastics = list()
        self.numprints = []
        #self.conc = []
        self.optionmenu_fluids = list()
        self.concs = list()
        
        #------- Bottom Frame  ----------#
        #variables assigned in Bottom Frame
      
        #widget bottom frame
        # Go Button
        self.btnGo = tkinter.Button(self.btm_frame, text='Calculate', width='15', command=self.btnGoClick)
        self.btnGo.config(stat='disabled')
        # Quit Button
        self.btnQuit = tkinter.Button(self.btm_frame,text='Quit',width='14',command=self.btnQuitClick)
        # Status bar
        self.status = tkinter.Label(self.btm_frame, width='50',fg="white",bg="gray", text='Status: Waiting')
        # Show Results
        self.btnRes = tkinter.Button(self.btm_frame, width='14',text="Show Results", fg="green",  command=self.open)
        self.btnRes.config(state='disabled')
        
        self.btnGo.grid(row=0,column=0,padx=15,sticky="w")
        self.status.grid(row=0,column=1,padx=25,sticky="ew")
        self.btnRes.grid(row=0,column=2,padx=25,sticky="ew")
        self.btnQuit.grid(row=0,column=3,padx=5,sticky="e")
        
        self.selrtype()
        
        if self.useInternalStd.get():
            self.selc1()
        
    #------- handle commands ----------#
    def open(self):
 #       start = 'start' + ' /D ' + self.OutPutFolder.get()
 #       os.system(start)
         folder = self.OutPutFolder.get()
         print('explorer folder: ', folder)
         wfolder = folder.replace('/','\\')
         subprocess.Popen('explorer "{0}"'.format(wfolder))
    
    def updateReadMode(self, value):
        if value == '0':
            self.readMode.set(True)
            self.readModeColor.set('chartreuse')
        elif value =='1':
            self.readMode.set(False)
            self.readModeColor.set('yellow')    
        self.readModeScale.config(troughcolor=self.readModeColor.get())
        self.getSupportedReaders()
        self.setLUT(None)
    
    def updateCurveMode(self,value):
        if value == '0':
            self.curveMode.set(True)
            self.curveModeColor.set('chartreuse')
        elif value =='1':
            self.curveMode.set(False)
            self.curveModeColor.set('cyan')
        self.curveModeScale.config(troughcolor=self.curveModeColor.get())
        self.getSupportedReaders()
        self.setLUT(None)

    def selc1(self):
     #   self.lblIn.config(text = self.inChoices[self.varRadio.get()] 
     #       + ' Input File Path')
     #   self.lblOut.config(text = self.inChoices[(self.varRadio.get()+1)%2] 
     #       + ' Output File Path')
        self.setLUT(None)
        if (self.useInternalStd.get()) == True :
            self.exclude16.set('True')
            self.toggleMatechEntry(state='disabled')
        else:
            self.exclude16.set('False')
            self.toggleMatechEntry(state='normal')
 #       print(' selc1 :: Internaldtd = ', str(self.useInternalStd.get()))
     
    def selc3(self):
        if (self.exclude16.get()) == False :
            self.useInternalStd.set('False')   
            self.entIn2.config(state='normal')
            self.entIn1.config(state='normal')
            self.btnIn1.config(state='normal')
            self.btnIn2.config(state='normal')
        
    def selrtype(self):
        if (self.runType.get() == self.RUNTYPES[0]):
            print(' -- Print mode -- ')
            self.entIn3.config(state='disabled')
            self.btnIn3.config(state='disabled')
            self.numSplitsPerSweep.config(state='disabled')
            if self.transferData:
                self.clearTransferFiles()
            self.sweepData = False
            self.transferData = False
            self.clearTransferFiles()
            self.setOutputFolder()
            self.c4.config(state='normal')
            self.toggleFTButtons(state='disabled')
            self.btnLoadFT.config(state='disabled')
            self.FTfolder.set(None)
 
        elif (self.runType.get() == self.RUNTYPES[1]):
            print(' -- Sweep mode -- ')
            self.entIn3.config(state='normal')
            self.btnIn3.config(state='normal')
            self.numSplitsPerSweep.config(state='normal')
            if self.transferData:
                self.clearTransferFiles()
            self.sweepData = True
            self.transferData = False
            self.setOutputFolder()
            self.c4.config(state='disabled')
            self.toggleFTButtons(state='disabled')
            self.btnLoadFT.config(state='disabled')
            self.FTfolder.set(None)
#            self.b1.config(state='disabled')
        elif (self.runType.get() == self.RUNTYPES[2]):
            print(' -- EPC FinalTest --')
            self.entIn3.config(state='disabled')
            self.btnIn3.config(state='disabled')
            self.numSplitsPerSweep.config(state='disabled')
            self.sweepData = False
            self.transferData = True
            self.btnLoadFT.config(state='normal')
            self.setOutputFolder()
            self.c4.config(state='disabled')
            if self.ReaderFileList:
                self.setupTransferFiles()
                self.populateFileFrame()
        else:
            print('Error, Unknown RunType')
        self.enforceSelectedLUT(self.selectedLUT.get())
        return

    def clearListFrame(self):
        try:
           for widget in self.listFrame.winfo_children():
               try:
                  widget.grid_remove()
               except:
                    print('can\'t remove basefilename !, i = ', widget) 
        except:
           print(' cwoool... no self.n')
        
    def toggleMatechEntry(self,state='normal'):
        self.entIn2.config(state=state)
        self.entIn1.config(state=state)
        self.btnIn1.config(state=state)
        self.btnIn2.config(state=state)
    
    def loadFTFolder(self):
        FTdir =  tkinter.filedialog.askdirectory(title = 'Select Final Test Folder')
        self.clearTransferFiles()
        if len(FTdir) > 0:
            if not (os.path.basename(FTdir) == 'FinalTest' or 'FT' in os.path.basename(FTdir)):
                message = 'Please select a Final Test folder.'
                messagebox.showinfo('Must be a Final Test folder', message)
                FTdir = tkinter.filedialog.askdirectory(title = 'Select Final Test Folder')
            self.FTfolder.set(FTdir)
            
            readerDir = FTdir+'/ReaderResults'
            self.wdir=readerDir
            self.tdir=readerDir
            self.ReaderFileList = [readerDir+'/'+r for r in os.listdir(readerDir) if not (r == 'destbarcodemap.csv' or 'rawvolumeoutput' in r.lower() or 
                                                                       'mergedat' in r.lower() or 'matech' in r.lower() or 'summary' in r.lower() or r.endswith('.xlsx')) and 
                                                                        (r.endswith('.csv') or r.endswith('.xls'))]
            self.loadedtransferfiles = EPCFT.getEPCData(False, wdir=FTdir)
            self.loadedtransferfiles.sort()
            if len(self.loadedtransferfiles) == len(self.ReaderFileList):
                for i in range(0,len(self.loadedtransferfiles)):
                    if len(self.loadedtransferfiles[i][-1]) != 5:
                        self.loadedtransferfiles[i][-1] = ['','','','',self.ReaderFileList[i]]
                    else:
                        for ii in range(0, len(self.loadedtransferfiles[i][-1][0:-1])):
                            self.loadedtransferfiles[i][-1][ii]=self.loadedtransferfiles[i][-1][ii].replace('./','/')
            else:
                messagebox.showerror("Warning", " Mismatch between TransferFiles and ReaderFiles.  Please manually link the files")
                self.loadedtransferfiles=[]
                self.setupTransferFiles()
            self.scrollb.grid(row=1, column=1, sticky='nsew')              
            self.updateMatechFiles()
            self.ReaderFileAttributes = {}
            self.populateFileFrame()
            self.PopulateTransferFiles()
            self.OutPutFolder.set(FTdir+'/Analysis')
            self.resetStatusBar()
            self.btnGo.config(state='normal')
            self.btnUpdateAttr.config(state='normal')    
        return

    def resetStatusBar(self):
        self.status.config(text='Waiting')
        self.status.config(bg='gray') 
        self.status.update()
        
    def updateMatechFiles(self):
        self.matechfiles = [self.wdir+'/'+f for f in os.listdir(self.wdir) if ('matech' in f.lower()) and (f.endswith('.csv') or f.endswith('.xls'))]
        if self.matechfiles:
#            self.useInternalStd.set(False)
#            self.exclude16.set(False)
#            self.toggleMatechEntry(state='normal')
            if len(self.matechfiles) == 1:
                self.MatechFileName1.set(self.matechfiles[0])
                self.MatechFileName2.set(self.matechfiles[0])
            elif len(self.matechfiles) > 1:
                self.matechfiles.sort()  
                try:
                    self.MatechFileName1.set(self.matechfiles[0])
                    self.MatechFileName2.set(self.matechfiles[-1])
                except:
                    print('Unable to load matechfiles to GUI')
                    return
        else:
#            self.useInternalStd.set(True)
#            self.exclude16.set(True)
#            self.toggleMatechEntry(state='disabled')
            print('Not enough Matech files')
            return

    def toggleFTButtons(self, state = 'disabled'):
        for child in self.center_FT_frame.winfo_children():
            child.configure(state=state)
    
    def btnOperationMode(self):
        if (self.operationMode.get() == 'CCT'):
            self.numSplitsPerSweep.config(state='disabled')
        else:
            self.numSplitsPerSweep.config(state='normal')
    
    def btnInBrowseClick1(self):             
        rFilepath = filedialog.askopenfilename(defaultextension='*', 
            initialdir=self.wdir, initialfile='', parent=self.center_frame, title='select a file') 
        self.MatechFileName1.set(rFilepath)
        self.wdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entIn1.get())
        
    def btnInBrowseClick2(self):             
        rFilepath = filedialog.askopenfilename(defaultextension='*', 
            initialdir=self.wdir, initialfile='', parent=self.center_frame, title='select a file') 
        self.MatechFileName2.set(rFilepath)
        self.wdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entIn2.get(), ' FilePath : ', rFilepath) 
        
    def btnSelectPickListFile(self):             
        rFilepath = filedialog.askopenfilename(defaultextension='*', 
            initialdir=self.wdir, initialfile='', parent=self.center_frame, title='select a file') 
        self.PickListFileName.set(rFilepath)
 #       self.wdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entIn3.get(), ' FilePath : ', rFilepath)     
    
    def btnSelectOutputFolder1(self):             
        rOutputFolder = filedialog.askdirectory(initialdir=self.wdir, parent=self.center_frame, title='select a folder') 
        self.OutPutFolder.set(rOutputFolder)
        
        
    def getSupportedLUTs(self):
        return self.ReaderDat.getCalibrations(reader = str(self.selectedReader.get()), UseInternalStd = self.useInternalStd.get(), MFG = self.curveMode.get(), Fluor=self.readMode.get())
        
    def getSupportedReaders(self):
        return self.ReaderDat.getSupportedReaders(UseInternalStd = self.useInternalStd.get(), MFG = self.curveMode.get(), Fluor=self.readMode.get())
    
    def setLUT(self, event):
        self.supportedLUTs = self.getSupportedLUTs()
        if self.selectedLUT.get() == 'UNSUPPORTED':
            try:
                self.selectedLUT.set( self.supportedLUTs[0])
            except:
                pass
        if self.selectedLUT.get() not in self.supportedLUTs:
            print('LUT ', self.selectedLUT.get(), ' not supported for reader ', self.selectedReader.get())
            try: 
                self.selectedLUT.set( self.supportedLUTs[0])
            except:
                self.selectedLUT.set('UNSUPPORTED')
#            self.selectedLUT.set( self.supportedLUTs[0])
        else:
            print('LUT ', self.selectedLUT.get(), ' is supported for reader ', self.selectedReader.get())
        print(' setLUT :: supportedLUTS = ', self.supportedLUTs )
        self.goUpdateAttributes()
        menu = self.m2['menu']
        menu.delete(0,'end')
        for lut in self.supportedLUTs:
            menu.add_command(label=lut,command=tkinter._setit(self.selectedLUT, lut,self.enforceSelectedLUT))
        print(' setLUT :: READER = ', self.selectedReader.get())
        
    def printLUT(self, event):
        print(self.selectedLUT.get())    
    
    def enforceSelectedLUT(self,lut):
        self.goUpdateAttributes()        
        
    def btnSelectFiles(self):
        self.ReaderFileList = filedialog.askopenfilenames(title='please select files you want to process',initialdir=self.wdir)
        self.clearTransferFiles()
 #       self.file_frame.grid(row=1,sticky="nsew")   
        if len(self.ReaderFileList) > 0 :
            self.wdir = os.path.dirname(self.ReaderFileList[0])
            self.tdir = os.path.dirname(self.ReaderFileList[0])
            self.setupTransferFiles()
        self.updateMatechFiles()
        self.setOutputFolder()            
        self.scrollb.grid(row=1, column=1, sticky='nsew')
      #  self.scrollb.pack()
        self.ReaderFileAttributes = {}
        self.populateFileFrame()
        self.btnGo.config(state='normal')
        self.btnUpdateAttr.config(state='normal')
        self.resetStatusBar()
        return
     
    def clearTransferFiles(self):
        self.clearListFrame()
        self.loadedtransferfiles = []
        self.tableIndex.set(0)
        self.surveyFileName.set('')
        self.printFileName.set('')
        self.MIPOutFileName.set('')
        
    def setupTransferFiles(self):
        if self.transferData:
            for i, r in enumerate (self.ReaderFileList):
                self.loadedtransferfiles.append(['0',['','','','',r]])
        else:
            print('Not in EPCFT mode')
            return

    def setOutputFolder(self):
        if self.transferData:
            if os.path.basename(self.wdir) == 'ReaderResults':
                self.OutPutFolder.set(os.path.dirname(self.wdir)+'/Analysis')
            else:
                self.OutPutFolder.set(self.wdir)
        else:
            self.OutPutFolder.set(self.wdir)   
        
    def goUpdateAttributes(self):
        if len(self.ReaderFileList) == 0 :
            return
        self.populateFileFrame()     
        
    def refreshTransferFilesUI(self):
        ind = self.tableIndex.get()
#        for i in range(0,len(self.loadedtransferfiles)):
        for ii in range(0, len(self.loadedtransferfiles[ind][-1])):
            if self.shortfile in self.loadedtransferfiles[ind][-1][ii]:
                selectedTransferfile = self.loadedtransferfiles[ind][-1]
                baseSelectedFiles = [(os.path.basename(b)).split('.csv')[0] for b in selectedTransferfile[0:-1] if 'MIPInfoEx' not in b]
                self.displayTransferFiles.set(baseSelectedFiles)
                self.transferFilesRadio = tkinter.Radiobutton(self.listFrame,
                                                              variable=self.tableIndex, text = self.displayTransferFiles.get(),
                                                              value=self.tableIndex.get(), 
                                                              command = self.PopulateTransferFiles) 
                self.transferFilesRadio.grid(row=self.n, column=6, sticky='w')
                self.tableIndex.set(ind)
#                self.transferFilesLabel = tkinter.Label(self.listFrame, text = self.displayTransferFiles.get())
#                self.transferFilesLabel.grid(row=self.n, column=7, padx=10, sticky='w')
                
#        self.rfileVar.set(self.ReaderFileList[int(self.tableIndex.get())])
    def checkFTstate(self):
        if self.ReaderFileList and self.transferData:
            self.toggleFTButtons(state='normal')
        else:
            self.toggleFTButtons(state='disabled')
    
    def populateFileFrame(self):
        self.plastics = ['Misc','384PPG', '384PPL', '1536LDVS', '384LDVS','6RES','384PPL-Plus','96TRK']
        self.fluids = ['Misc','DMSO','SP','BP/GP','AQ','CP-PEG','CP-HEPES','CP-MPD']
        headers=['Filename','Plastic','Fluid','Conc','PrintNr','LUTDate','TransferFiles']
        self.checkFTstate()
        for i,h in enumerate(headers):
            self.fileFrameHeader1 = tkinter.Label(self.listFrame, text=h)
            
            self.fileFrameHeader1.grid(row=0,pady=5,padx=5,column=i)
        self.n = 1    
        for ind,file in enumerate(self.ReaderFileList):
            self.tableIndex.set(ind)
            thisfile = {}
            self.shortfile = os.path.basename(file).lstrip()
            if self.shortfile in self.ReaderFileAttributes:
                attribute = self.ReaderFileAttributes[self.shortfile]
                plastic = attribute['Plastic'].get()
                fluidclass = attribute['Fluid'].get()
                conc = attribute['Conc'].get()
                printnr = attribute['Print'].get()
            else:
                plastic = 'Misc'
                fluidclass = 'Misc'
                conc = 'NA'
                printnr = '1'
            _lut = self.selectedLUT.get()
            if self.operationMode.get() == 'CCT' :
                 barcode = os.path.basename(file).split('PlateBC_')[1].split('_Labcyte')[0]
                 barcode = barcode.replace('-','_')
                 if len(barcode.split('_')) == 2 :
                    bc1 = barcode.split('_')[0]
                    bc2 = barcode.split('_')[1]
                    # plastic is penultimate entry in first part
                    plastic = processCCT.plasticdecode(bc1[-2])
                    try:
                        _lut = processCCT.lutdecode(bc2[-3:])
                    except:
                        print('invalid filename found when looking for _lut')
                        _lut = self.selectedLUT.get()
                    fluidclass = processCCT.fluiddecode(bc2[:-6])
                    conc = bc2[-6:-3]
                    try:
                        cc = int(conc)
                        conc = ( cc//10) *10
                    except:
                        conc = conc
                    printnr = bc2[-4:-3]   
                    if self.n == 1 :
                        self.EchoSN.delete(0,tkinter.END)
                        self.EchoSN.insert(0,'E' + bc1[:4])
                    if ( self.echoType.get() == 'E525' ):
                        self.nomVol.set(100)
                    else:
                        self.nomVol.set(50)
                        
            if 'FinalTest' in self.FTfolder.get():               
                plateType = self.FTfolder.get().split('FinalTest')[0].split('/')[-2]
                plastic = plateType.split('_')[0]
                fluidclass=plateType.split('_')[-1]
                if fluidclass == 'DMSO2':
                    fluidclass = 'DMSO'
                conc = os.path.basename(file).split('__')[0]

      #  while self.n <= int(len(self.ReaderFileList)):
            self.basefilename = tkinter.Label(self.listFrame, width='40', text = os.path.basename(file), anchor='w',justify="left")
            self.basefilename.grid(row=self.n, column=0,padx=5,pady=5)

            var = tkinter.StringVar()
            var.set(plastic)
            menu = tkinter.OptionMenu(self.listFrame, var, *self.plastics)
            menu.config(width=10)
            menu.grid(row=self.n, column=1, padx=10,pady=5)
            
            self.optionmenu_plastics.append((menu, var))
            thisfile['Plastic'] = var 
            
            var2 = tkinter.StringVar()
            menu2 = tkinter.OptionMenu(self.listFrame, var2, *self.fluids)
            menu2.config(width=10)
            menu2.grid(row=self.n, column=2,padx=10,pady=5)
            var2.set(fluidclass)
            self.optionmenu_fluids.append((menu2, var2))
            thisfile['Fluid'] = var2 

            self.conc = tkinter.Entry(self.listFrame,width=6)
            self.conc.grid(row=self.n,column=3,padx=10)
            self.conc.insert(0,conc) 
            self.concs.append(self.conc)
            thisfile['Conc'] = self.conc
            
            self.printnr = tkinter.Entry(self.listFrame,width=6)
            self.printnr.grid(row=self.n, column=4, padx=10)
            self.printnr.insert(0,printnr)
            self.numprints.append(self.printnr)
            thisfile['Print'] = self.printnr

            thislut = tkinter.StringVar()
#            print(' file: ', file, ' lut: ', _lut)
            if _lut in self.supportedLUTs:
                thislut.set(_lut)
            else:
                thislut.set(self.selectedLUT.get())
            lutdate = tkinter.OptionMenu(self.listFrame, thislut,*self.supportedLUTs)
            lutdate.config(width=8)
            lutdate.grid(row=self.n, column=5, padx=10)
            thisfile['LUT'] = thislut
            
            if self.transferData:
                self.refreshTransferFilesUI()
                self.tableIndex.set(0)
                thisfile['TransferFiles'] = self.loadedtransferfiles[self.tableIndex.get()]
            else:
                thisfile['TransferFils'] = []

                    
            self.n += 1
            self.ReaderFileAttributes[os.path.basename(file).lstrip()] = thisfile
                                      
    #    print(self.canvas.children())                              
        
    #    print(self.listFrame.winfo_children())      
    
    def PopulateTransferFiles(self):
        ind = self.tableIndex.get()
        try:
            self.surveyFileName.set([s for s in self.loadedtransferfiles[ind][-1] if 'platesurvey.csv' in s][0])
        except:
            self.surveyFileName.set('')
        try:
            self.printFileName.set([s for s in self.loadedtransferfiles[ind][-1] if 'print.csv' in s][0])
        except:            
            self.printFileName.set('')
        try:
            self.MIPOutFileName.set([s for s in self.loadedtransferfiles[ind][-1] if 'MIPOutput.csv' in s][0])
        except:
            self.MIPOutFileName.set('')
    
        return self.surveyFileName.get(); self.printFileName.get(); self.MIPOutFileName.get()

    def btnSurveyClicked(self):
        originalIndex = self.tableIndex.get()
        rFilepath = filedialog.askopenfilename(defaultextension='*platesurvey.csv', filetypes = (('csv files', 'platesurvey.csv'),),
            initialdir=self.tdir, initialfile='', parent=self.center_FT_frame, title='Select Survey file') 
        self.surveyFileName.set(rFilepath)
        self.tdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entFT1.get())
        getIndex = [i for i,f in enumerate(self.loadedtransferfiles[self.tableIndex.get()][-1]) if 'platesurvey' in f]
        if len(getIndex)>0:
            self.loadedtransferfiles[self.tableIndex.get()][-1][getIndex[0]] = self.surveyFileName.get()
        else:
            self.loadedtransferfiles[self.tableIndex.get()][-1][0]=self.surveyFileName.get()
        self.goUpdateAttributes()
        self.tableIndex.set(originalIndex)
        return self.entFT1.get(), self.displayTransferFiles.get()
    
    def btnPrintClicked(self):
        originalIndex = self.tableIndex.get()
        rFilepath = filedialog.askopenfilename(defaultextension='*print.csv', filetypes = (('csv files', 'print.csv'),),
            initialdir=self.tdir, initialfile='', parent=self.center_FT_frame, title='Select Print file') 
        self.printFileName.set(rFilepath)
        self.tdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entFT2.get())
        getIndex = [i for i,f in enumerate(self.loadedtransferfiles[self.tableIndex.get()][-1]) if 'print' in f]
        if len(getIndex)>0:
            self.loadedtransferfiles[self.tableIndex.get()][-1][getIndex[0]] = self.printFileName.get()
        else:
            self.loadedtransferfiles[self.tableIndex.get()][-1][1]=self.printFileName.get()
        self.goUpdateAttributes()
        self.tableIndex.set(originalIndex)
        return self.entFT2.get(), self.displayTransferFiles.get()

    def btnMIPOutClicked(self):
        originalIndex = self.tableIndex.get()
        rFilepath = filedialog.askopenfilename(defaultextension='*MIPOutput.csv', filetypes = (('csv files', 'MIPOutput.csv'),),
            initialdir=self.tdir, initialfile='', parent=self.center_FT_frame, title='Select MIPOutput file') 
        self.MIPOutFileName.set(rFilepath)
        self.tdir = os.path.dirname(rFilepath)
        print(' Filename : ' ,  self.entFT3.get())
        getIndex = [i for i,f in enumerate(self.loadedtransferfiles[self.tableIndex.get()][-1]) if 'MIPOutput' in f]
        if len(getIndex)>0:
            self.loadedtransferfiles[self.tableIndex.get()][-1][getIndex[0]] = self.MIPOutFileName.get()
        else:
            self.loadedtransferfiles[self.tableIndex.get()][-1][2]=self.MIPOutFileName.get()
        self.goUpdateAttributes()
        self.tableIndex.set(originalIndex)
        return self.entFT3.get(), self.displayTransferFiles.get()
                                     
    def onFrameConfigure(self,canvas):
        '''Reset the scroll region to encompass the inner frame'''
        canvas.configure(scrollregion=canvas.bbox("all"))                                  
        
    def AuxscrollFunction(self,event):
        #You need to set a max size for frameTwo. Otherwise, it will grow as needed, and scrollbar do not act
        self.canvas.configure(scrollregion=self.canvas.bbox("all"),height=self.listframeheight,width=self.listframewidth)
        return
        
    def btnGoClick(self):  
        if len(self.ReaderFileList) == 0:
            messagebox.showerror("Error", "Please specify some reader files to parse")
            self.setfail()
            return
        try:
            self.ReaderDat.setStdCurve(lutdate=self.selectedLUT.get(),reader=self.selectedReader.get(),UseInternalStd=self.useInternalStd.get())
        except:
            messagebox.showerror('Error in Standard Curve','Incorrect Standard Curve combination.  Check if Std Curve is supported')
            self.setfail()
            return
        
        if not(self.useInternalStd.get()):
            if not os.path.exists(self.MatechFileName1.get()):
                messagebox.showerror("Error", "Matech File 1 does not exist")
                self.setfail()
                return
            if not os.path.exists(self.MatechFileName2.get()):
                messagebox.showerror("Error","Matech File 2 does not exist")
                self.setfail()
                return
            try:
                self.ReaderDat.setPMTCorr(f1=self.MatechFileName1.get(),f2=self.MatechFileName2.get())
            except:
                messagebox.showerror('Error','Problem reading Matech files, double check selected reader')
                self.setfail()
                return
            
        if os.path.basename(self.OutPutFolder.get()) == 'Analysis':
            if not os.path.exists(self.OutPutFolder.get()):
                os.makedirs(self.OutPutFolder.get())
        elif not os.path.isdir(self.OutPutFolder.get()   ):
            messagebox.showerror('Error','Output folder is not a valid folder -- Folder needs to exist')
            self.setfail()
            return
        self.status.config(text='Busy')
        self.status.config(bg='red') 
        self.status.update()
        self.prefix =  datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S') 
        self.outfile = self.OutPutFolder.get()+'\\'+self.prefix+'_ReaderDat_DataSummary.xlsx'
        if self.sweepData:
            ret = self.parseSweep()
        elif self.transferData:
            ret = self.runEPCFT()
        else:    
            ret = self.parsePrint()
       
        if ret == 0 :
            self.status.config(text='Done')
            self.status.config(bg='green') 
            self.btnRes.config(state='normal')
        else:
            self.setfail()
            
            
        print('Done.')
        return 
    
    def setfail(self):
        self.status.config(text='Error')
        self.status.config(bg='tomato')
        return
    
    def parseSweep(self):
        
        AllDatDf = pandas.DataFrame()
        destrows = 16
          
        for readerfile in self.ReaderFileList:  
            
            print( ' ' )
            print(' ===== processing file ', readerfile, ' ===== ')
            self.shortfile = os.path.basename(readerfile).lstrip()
            attributes= self.ReaderFileAttributes[self.shortfile]
            self.processES.UseInternalStd  = self.useInternalStd.get()
            self.processES.reader=self.selectedReader.get();
            self.processES.lutdate=attributes['LUT'].get()
            self.processES.applyPMTCorr = self.applyPMTCorr.get()
            self.processES.exclude16 = self.exclude16.get()
            
            if not self.goodReaderCombo(attributes):
                return -1
            
 #           self.processES.readPickListFile()
 #           self.processES.genPickListMtrxFromArray(self.processES.xrange_cct)
 #           self.processES.genPickListMtrxFromArray(self.processES.xrange_tswap)
 
            if self.PickListFileName.get() == '' :
                if self.operationMode.get() == 'CCT':
                        self.processES.genPickListMtrxFromArray(self.processES.xrange_cct)
                else:
                    if self.useInternalStd.get():
                        self.processES.genPickListMtrxFromArray(self.processES.xrange_tswap_IntStd)
                    elif self.exclude16.get():
                        self.processES.genPickListMtrxFromArray(self.processES.xrange_tswap_IntStd)
                    else:    
                        self.processES.genPickListMtrxFromArray(self.processES.xrange_tswap)
            else:
                self.processES.readPickListFile(self.PickListFileName.get())
            
                    
    #            self.processES.genPickListMtrxFromArray(self.processES.xrange_tswap)
            try:    
                fname = self.prefix + '_' + self.shortfile.split(self.processES.csv_or_xls())[0]  + '_flood'
                if self.useInternalStd.get():
                    eSweepDict = self.processES.processEjectSweepRawReader([readerfile], 
                                                                           filename = fname , outdir = self.OutPutFolder.get())
                else:
                    eSweepDict = self.processES.processEjectSweepRawReader([readerfile, self.MatechFileName1.get(), self.MatechFileName2.get()], 
                                                                            filename=fname, outdir=self.OutPutFolder.get())
            except AssertionError as error:
                messagebox.showerror('Error', error)
                return -1
            if len(eSweepDict) == 0:
                messagebox.showerror("Error", " PMT Correction factor likely too high")
                return
            ## get from  make attribute
            try:
                numsplitspersweep = int(self.numSplitsPerSweep.get())
            except:
                messagebox.showerror("Error", "# SweepTestsPerPlate needs to be an integer")
                return -1
            ## TODO make sure numsplitspersweep is divider of nr of rows . 
            if destrows%numsplitspersweep:
                messagebox.showerror("Error", "The # of sweeps needs to be an integer divider of the numbers of rows in the dest plate. For example: 1,2,4,8 or 16 ")
                return -1
            
            ## TODO add logic:
            # if DMSO and CCT:  numsplitspersweep = 4
            # if SP and CCT: numsplitspersweep = 1
            concs = attributes['Conc'].get()
            if self.operationMode.get() == 'CCT' :
                if 'SP' in attributes['Fluid'].get():
                    numsplitspersweep = 2
                    concs = [ '14','200']
                if 'DMSO' in attributes['Fluid'].get():
                    numsplitspersweep = 4
                    concs = [ '70','90','100','80']
            platetype = self.determinePlateType(plastic = attributes['Plastic'].get(), fluid = attributes['Fluid'].get(), conc = attributes['Conc'].get())
            # todo platetyp constructor based on fluid and plastic attribute
            if numsplitspersweep > 1:
                ncols = 2
                nrows = (numsplitspersweep+1)//2 
            else:
                ncols = 1
                nrows = 1
            fig, ax = plt.subplots(nrows=nrows, ncols=ncols, figsize=(9*ncols,3*nrows ))           
            gs = processEjectSweep.gridspec.GridSpec(nrows,3*ncols);        
            gs.update(wspace=0.3,hspace=0.4)
            irow = 0
            icol = 0   
            pltcnt = 0
          
            self.verbose=True
            for ff, res in eSweepDict.items():
                voldat = res['VolDat']
                pmtcor = res['PMTCorr']
                for k in range(0,numsplitspersweep):
                    startRow = 0+k*(destrows//numsplitspersweep)
                    stopRow = (k+1)*(destrows//numsplitspersweep)
                    print(' --- k = ', k, ' -- startrow: ', startRow, ' stopRow: ', stopRow)
                    try:
                        y, ye, x, cv = self.processES.calcStats(voldat, startRow, stopRow) 
                    except AssertionError as error:
                        messagebox.showerror('Error', error)
                        return -1
                        
                    title = os.path.basename(ff).split(self.processES.csv_or_xls())[0]
                    if numsplitspersweep > 1 :
                        title += '_' + str(k)
                    if self.verbose: 
                        print( ' irow: ', irow, ' icol: ', icol, ' pltcnt : ', pltcnt, ' t: ', (pltcnt+1)%nrows)
                    self.processES.platesettings[platetype]['nominalvolume'] = float(self.nomVol.get())
                    A = self.processES.plotBowls([x,y,ye,cv], ax=[gs[irow,icol:icol+2],gs[irow,icol+2]], plotname=title, 
                                                    readerPMT=pmtcor, thisplatetype=self.processES.platesettings[platetype], basedir= self.OutPutFolder.get())
                   # self.processES.setYRange(platetype)
                  #  gs[irow,icol:icol+2].set_ylim(self.processES.ylo,self.processES.yhi)
                    B = pandas.DataFrame.from_dict(A, orient='index').T
                    B['Platetype'] = platetype
                    try:
                        if self.operationMode.get() == 'CCT' :
                            B['Concentration'] = concs[k]
                        else:
                            B['Concentration'] = concs
                    except:
                        B['Concentration'] = concs
                    B['Plastic'] = attributes['Plastic'].get()
                    B['Sweep'] = k
                    B['PrintNr'] = attributes['Print'].get()
                    B['Fluid'] = attributes['Fluid'].get()
                    AllDatDf = AllDatDf.append(B)
       
                    if ((pltcnt+1)%ncols) == 0 :
                        icol = 0
                        irow += 1
                    else:
                        icol += 3
                    pltcnt += 1    
                 
            try:                          
                ofile = self.shortfile.split(self.processES.csv_or_xls())[0]
                print(' -- saving to file :: ', ofile, ' in folder ', self.OutPutFolder.get(), ' -- ')
                file = self.OutPutFolder.get() + "\\" + self.prefix + '_' + ofile + '_bowl.png'
                fig.savefig(file,bbox_inches='tight')
                
            except Exception as error:
                messagebox.showerror('Error', 'Can\'t write to file ' + ofile + '_bowl.png')
    

        AllDatDf['SN'] = self.echoType.get() 
        print('AllDat :: ', AllDatDf)
        print('AllDatDF.columns:: ', AllDatDf.columns)
        dfcols = ['Name', 'Platetype', 'Concentration', 'Sweep','PrintNr','T2T adjust', 'CF adjust',  
              'wallOfDeath', 'BowlLeft', 'BowlCenter', 'BowlRight',    'BowlVol_Center',   'BowlVol_Std',  'BowlVol_Mean', 'BowlVol_0dB',  'ReaderPMT'  ]
            
       
        AllDatDf = AllDatDf[['SN','Platetype', 'Plastic','Fluid','Concentration', 'Sweep','PrintNr','T2T adjust', 'CF adjust','wallOfDeath', 
                                 'BowlLeft', 'BowlCenter', 'BowlRight',    'BowlVol_Center',    'BowlVol_0dB',  'ReaderPMT' ,'Name']]
        print('AllDat :: ', AllDatDf)     
        fbase = self.OutPutFolder.get() + './' + self.EchoSN.get() + '_' + self.prefix         
        try:    
 #TODO make multi-index ; xl file and reduce the columns,                       
            AllDatDf.to_csv( fbase + '_BowlAnalysis.csv', index=False) 
        except Exception as error:
            messagebox.showerror('Error','Can\'t write to file ' +  fbase + '_BowlAnalysis.csv')      
            
        R = AllDatDf.set_index(['SN','Fluid','Plastic','Concentration'])    
        R = R.sortlevel(['Fluid','Plastic','Concentration'], sort_remaining=False ) 
        
        try:
            self.writeXLS(dataframe=R,xlfilename= fbase + '_BowlAnalysis.xlsx')
        except Exception as error:
            messagebox.showerror('Error','Can\'t write to file ' + fbase + '_BowlAnalysis.xlsx')
            
        return 0 
    
    def outputRawVol(self, rfile, cal):
        dc = pandas.DataFrame(cal)
        dc = dc.round(2)
        dc.index=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
        dc.columns = dc.columns + 1
        rout = os.path.basename(rfile).replace('.csv','').replace('.xls','').replace('PlateBC_','')
        sheet_name = rout[0:23]+'_'+rout[-6:]
#        if not os.path.exists(self.outfile):
#            workbook = openpyxl.Workbook()
#            workbook.save(self.outfile)
        self.writeWorksheet(sheet_name,dc,self.outfile,applyCond=True)

    def writeWorksheet(self,sheet_name, dataframe, outfile, applyCond=False):
        if not os.path.exists(outfile):
            workbook = openpyxl.Workbook()
            workbook.save(outfile)
        workbook = openpyxl.load_workbook(outfile)
        writer = pandas.ExcelWriter(outfile, engine='openpyxl')
        writer.book = workbook
        dataframe.to_excel(writer,sheet_name=sheet_name)
        try:
            std = workbook['Sheet']
            workbook.remove(std)
        except:
            pass
        worksheet = workbook[sheet_name]
        if applyCond:
            mean = dataframe.values.mean()
            self.applyFormat(worksheet,mean)
        workbook.active = len(workbook.sheetnames)-1
        writer.save()
        writer.close()
        
    def applyFormat(self,worksheet,mean):
        rule = ColorScaleRule(start_type='min', start_value=0.6*mean, start_color='247CBD',
#                      mid_type='percentile', mid_value=50, mid_color='e3ff03',
                      end_type='max', end_value=1.4*mean, end_color='ffc7ce')
        worksheet.conditional_formatting.add('B2:Y17', rule)
            
    def writeXLS(self,dataframe, xlfilename):
        writer = pandas.ExcelWriter(xlfilename,engine='xlsxwriter')
        dataframe.to_excel(writer, sheet_name='report')
        workbook = writer.book
        worksheet = writer.sheets['report']
        dec_fmt = workbook.add_format({'num_format': '0.0', 'align':'center'})
        pmt_fmt = workbook.add_format({'num_format': '0.00', 'align':'center'})
        pass_fmt = workbook.add_format({'align': 'center'})
        
        cols = ['BowlVol_0dB','BowlVol_Center']
        dataframe[cols] = dataframe[cols].apply(pandas.to_numeric, errors='coerce')
 
        print(dataframe.dtypes)
    
        #    worksheet.set_column('F:H', 9, dec_fmt)
        worksheet.set_column('D:D',15,pass_fmt)
        worksheet.set_column('E:E',15)
        worksheet.set_column('F:J',15, dec_fmt)
        worksheet.set_column('K:M',10, pmt_fmt)
        worksheet.set_column('N:O',15, dec_fmt)
        worksheet.set_column('P:P',9, pmt_fmt)
        worksheet.set_column('Q:Q',80, pass_fmt)
       
        
        lines = len(dataframe)+1
        
        footformat = workbook.add_format({'font_size':10})
        footformat2 = workbook.add_format({'font_size':10, 'italic':True})
        ftst = lines+2
        ftcol = 0       
        worksheet.write(ftst,ftcol, 'Reader', footformat)
        worksheet.write(ftst,ftcol+1, self.selectedReader.get(), footformat2)
#        worksheet.write(ftst+1,ftcol, 'LUT',footformat)
#        worksheet.write(ftst+1,ftcol+1, self.selectedLUT.get(),footformat2)
   #     worksheet.write(ftst+1,ftcol, 'ProcDate',footformat)
   #     worksheet.write(ftst+1,ftcol+1, dataframe.date,footformat2)
        worksheet.write(ftst+1,ftcol,'Date',footformat)  
        worksheet.write(ftst+1,ftcol+1, time.strftime('%Y%m%d'), footformat2 )   
        worksheet.write(ftst+2,ftcol,'UsePMTCorr', footformat)
        worksheet.write(ftst+2,ftcol+1, self.applyPMTCorr.get(), footformat2) 
        worksheet.write(ftst+3,ftcol,'ExcludeCol16', footformat)
        worksheet.write(ftst+3,ftcol+1, self.exclude16.get(), footformat2) 
              
        writer.save()
        return  

    def runEPCFT(self):
        if not self.ReaderFileList:
            print(' **** ERROR **** : no ReaderFiles found')
            return
        
        allall = []
        print(self.loadedtransferfiles)
        for datalist in self.loadedtransferfiles:
            readerfile = datalist[1][-1]
            filelist = datalist[1]
            if "" in filelist[0:-3]:
                tkinter.messagebox.showerror('EPCFT Processing Error',' Some Transfer Files for '+os.path.basename(readerfile)+' are not defined.')
                return
            else:
                print('processing ', readerfile)
                self.shortfile = os.path.basename(readerfile).lstrip()
                attributes= self.ReaderFileAttributes[self.shortfile]
                plastic = attributes['Plastic'].get()
                fluid = attributes['Fluid'].get()
                conc = attributes['Conc'].get()
                if 'NA' in conc:
                    conc = self.shortfile
                if datalist[0] != conc:
                    datalist[0] = conc
                if '384PPL' in plastic and 'DMSO' in fluid:
                    fluid = 'DMSO2'
                if 'DMSO' in fluid:
                    platetype = plastic+'_'+fluid
                else:
                    platetype = plastic+'_AQ_'+fluid
                
                m_transfer = EPCFT.processTransferFiles(filelist)
                
                self.shortfile = os.path.basename(readerfile)
                attributes= self.ReaderFileAttributes[self.shortfile]
                self.goodReaderCombo(attributes)
                try:
                    cal, stats = self.processReaderFile(readerfile, attributes['LUT'].get())
                    self.outputRawVol(self.shortfile,cal)
                except:
                    tkinter.messagebox.showerror('Unable to process Reader Curve','Possible Std Curve issue')
                    return
                
                dc_df = self.ReaderDat.readertoDF(cal, conc)
                m_all = pandas.merge(dc_df,m_transfer,on='DestWell')
    #            m_all = m_all.infer_objects()
                m_all = m_all.convert_objects(convert_numeric=True)
                try:
                    m_all['PowerState'] = m_all['PowerState'].apply(hex)
                except:
                    print('No MIP files')
                d = {}
                d['Conc']=conc
                d['Reader']=cal
                d['Stats']=stats
                d['Merged']=m_all
                
    #            d = EPCFT.processTransferData(datalist,self.ReaderDat,pattern=0xFFFF)
                if not (d) :
                    print(' Invalid return in processEPCData while processing concentration ', datalist[0])
                    tkinter.messagebox.showerror('EPCFT Processing Error',' Invalid return in processEPCData while processing.  Check parameters, i.e. Std Curve, etc.')
                    return  
                allall.append(d) 
                # check if RT MIP enabled
                if (len(d['Merged'].columns) > 60 ):
                    EPCFT.genMergeSumPlots(d['Merged'],str(datalist[0]) + '__' + platetype + '__' + self.prefix, foldername = self.OutPutFolder.get())
                else:
                    EPCFT.genMergeSumPlots(d['Merged'],str(datalist[0]) + '__' + platetype + '__' + self.prefix ,  ['ReaderVol','ReaderVol','ReaderVol', 'ArbAmplitude (Volts)'],
                                                                       ['FluidComposition','FluidThickness (mm)','ArbAmplitude (Volts)', 'FluidThickness (mm)'], foldername = self.OutPutFolder.get())
        mergedat = allall[0]['Merged']
        for k in range(1,len(allall)):
            mergedat = mergedat.append(allall[k]['Merged'])
        
        self.writeWorksheet('MergeDat',mergedat,self.outfile)
#        mergedat.to_excel(self.OutPutFolder.get()+'/MergeDat__' + platetype + '__' + self.prefix + '.xls')   
        
        allstats = {}    
        for k in range(0,len(allall)):
            allstats[allall[k]['Conc']]=allall[k]['Stats']
        df = pandas.DataFrame(allstats).T  
        df.columns = self.ReaderDat.PRINTSTATS_FIELDS
#        df.sort_index().to_csv(self.OutPutFolder.get()+'/DataSummary__' + platetype + '__' + self.prefix + '.csv',  float_format='%.2f', date_format='%Y-%m-%d_%H:%M:%S')
        print(' ---- generating matrix plots ----')
        print( df.sort_index() ) 
        self.parsePrint()
        return 0
    
    def processReaderFile(self, readerfile, LUT):
        print( ' ' )
        print(' ===== processing file ', readerfile, ' ===== ')
        self.ReaderDat.setStdCurve(lutdate=LUT,reader=self.selectedReader.get(),UseInternalStd=self.useInternalStd.get())
 #           print(' -- use Internal Std -- ', self.useInternalStd.get())
        if not(self.useInternalStd.get()):
  #              self.ReaderDat.setPMTCorr(f1=self.MatechFileName1.get(),f2=self.MatechFileName2.get())
            try:
                self.ReaderDat.setPMTCorr(f1=self.MatechFileName1.get(),f2=self.MatechFileName2.get())
            except:
                messagebox.showerror('Error','Problem reading Matech files, double check selected reader')
                self.setfail()
                return
        self.ReaderDat.applyPMTCorr = self.applyPMTCorr.get()
        self.ReaderDat.exclude16 = self.exclude16.get()
        raw = self.ReaderDat.processReaderFile(readerfile)
        cal = self.ReaderDat.applyCalCurve(raw)            
        if not self.ReaderDat.readerdatAvailable :
            messagebox.showerror("Error:: ", "PMT correction of " +str(self.ReaderDat.StdCurvePMTCorr)+" too high ? ( Tolerance : " + str(self.ReaderDat.PMTCorrectionTolerance) + " )")
            return -1
        stats = self.ReaderDat.getPrintStats(cal)
        return cal, stats

    def printSummary(self):
        prefix = self.prefix  
        outfilename = prefix + '_ReaderDat'
        fulloutfilename = self.OutPutFolder.get() + '/' + outfilename    
        ncols = 4
        nrows = (len(self.ReaderFileList)+(ncols-1))//ncols
        if (nrows == 1 ):
            nrows += 1 
        fig, ax = plt.subplots(figsize=(18*ncols,10*nrows),nrows=nrows,ncols=ncols);
        row = 0
        col = 0
        for i in range(1,nrows):
            for j in range(1, ncols):
                ax[i][j].axis('off')
 #       allstats = []
        
        AllStats = pandas.DataFrame()
        
        for readerfile in self.ReaderFileList:  
            self.shortfile = os.path.basename(readerfile)
            attributes= self.ReaderFileAttributes[self.shortfile]
            self.goodReaderCombo(attributes)
            cal, stats = self.processReaderFile(readerfile, attributes['LUT'].get())
            if (self.rawVolOutput.get()) and not self.transferData:
                self.outputRawVol(self.shortfile,cal)
            # need to plot using named tuple stats
            if self.ReaderDat.readerdatAvailable:
                self.ReaderDat.plotReaderDat(cal,stats, ax=ax[row][col], title=os.path.basename(readerfile))
        
            # now calculate real stats. Note that CCT returns dataframe 
 #           print(' Mode :: ', self.operationMode.get())
            if self.operationMode.get() == 'CCT' :
                if (('DMSO' in attributes['Fluid'].get()) and  not ( '96TR' in attributes['Plastic'].get())):
                    stats = processCCT.processDMSOUCP(cal,self.ReaderDat)
                elif 'SP' in attributes['Fluid'].get() :
                    if ( self.echoType.get() == 'E525' ):
                        stats = processCCT.process525UCP(cal, self.ReaderDat)
                    else:    
                        stats = processCCT.processSPUCP(cal,self.ReaderDat)
                else:
                    if ( self.echoType.get() == 'E525' ):
                        stats = processCCT.process525PlusUCP(cal,self.ReaderDat,attributes['Conc'].get())
                    else :
                        stats = processCCT.processUCP(cal,self.ReaderDat)
                        stats['Concentration'] = attributes['Conc'].get()
            else:    
                statdict = stats._asdict()
                stats = pandas.DataFrame(statdict,index = ['0'])
#                print(' === stats === ')
#                print(stats)
                stats['Concentration'] = attributes['Conc'].get()
            
            stats['File'] = self.shortfile    
            stats['Plastic'] = attributes['Plastic'].get()
            stats['Fluid'] = attributes['Fluid'].get()
            stats['EchoSN'] = self.echoType.get()  
            stats['PrintNr'] = self.printnr.get()
            stats['LUT'] = attributes['LUT'].get()
            
            AllStats = AllStats.append(stats)
            
            if col < (ncols-1) :
                col +=1
            else:
                col = 0
                row +=1
        plt.savefig(fulloutfilename + '.png')
    #    AllStats = pandas.DataFrame(allstats)
 #       print(' **** ALLSTATS **** ')
 #       print(AllStats)
        CC = ['File','EchoSN','Plastic', 'Fluid','N','Mean','RawCV','CV','Empties','Outlier','Reliability','PMTCorr','LUT']
        #AllStats_csv = AllStats[CC]
        self.writeWorksheet('ReaderDat_Summary',AllStats[CC],fulloutfilename+'_DataSummary.xlsx')
#        AllStats[CC].to_csv(fulloutfilename +'_DataSummary.csv', index=False)
        return AllStats

    def parsePrint(self):
        prefix = self.prefix
        AllStats = self.printSummary()            
        if (AllStats.empty):
            print(' no valid files found for echo ', self.EchoSN.get() )
            return -1
        AllStats.Reader = self.selectedReader.get() 
        AllStats.applyPMTCorr = self.applyPMTCorr.get()
        if (len(self.shortfile.split('_')) > 4 ):
            AllStats.date = self.shortfile.split('_')[4]            
        else:
            AllStats.date = '160619'
#        AllStats.LUT =   self.selectedLUT.get()
        AllStats.SN = self.EchoSN.get()
     #   allstats['Plastic'] = plastic
        nomvol = float(self.nomVol.get())
        AllStats.nomvol = nomvol    
        AllStats['SN'] = self.EchoSN.get()
        AllStats['Status'] = 'PASS'
        crit= {}        

        crit['PMTCorrpct'] = 5   
        if (self.passfailCriteria.get()  == 'MFG' ):
           print(' using MFG criteria ')
           crit['dmsoaccuracy'] = 0.05
           crit['dmsocvpct'] = 5
           crit['accuracy'] = 0.08
           crit['cvpct'] = 6
           crit['cpaccuracy'] = 1
           crit['cpcvpct'] = 6
           crit['empties'] = 0
           crit['reliabilities'] = 0
           crit['pmtcorr'] = 0.03 
        elif (self.passfailCriteria.get()  == 'SER' ):
            print(' using Service criteria ')
            crit['accuracy'] = 0.08
            crit['dmsoaccuracy'] = crit['accuracy']
            crit['cpaccuracy'] = 1
            crit['cvpct'] = 6
            crit['dmsocvpct'] = crit['cvpct']
            crit['cpcvpct'] = crit['cvpct']
            crit['empties'] = 0
            crit['reliabilities'] = 1
            crit['pmtcorr'] = 0.10 
        else:
            print(' using CUST criteria ')
            crit['accuracy'] = 0.1
            crit['dmsoaccuracy'] = crit['accuracy']
            crit['cpaccuracy'] = 1
            crit['cvpct'] = 8
            crit['dmsocvpct'] = crit['cvpct']
            crit['cpcvpct'] = crit['cvpct']
            crit['reliabilities'] = 1
            crit['empties'] = 0
            crit['pmtcorr'] = 0.10


#        print(AllStats)
#        print(AllStats.dtypes)
        # want Concentration as string, as I am going to do string comparison next
        # strings are needed because of "NA" option ( yes it could be NaN too )
        AllStats['Concentration'] = AllStats['Concentration'].astype(str)
 #       print(AllStats.dtypes)
        AllStats['Status'][(AllStats['Fluid'] == 'DMSO') & (AllStats['RawCV']> crit['dmsocvpct'])] = 'FAIL'
        AllStats['Status'][(AllStats['Fluid'] == 'DMSO') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['dmsoaccuracy']) ] = 'FAIL'   
        if self.operationMode.get() == 'CCT' :
            AllStats['Status'][(AllStats['Fluid'] != 'BP/GP') & (AllStats['Concentration'] == '20') & (AllStats['RawCV']> crit['cpcvpct'])] = 'FAIL'
            AllStats['Status'][(AllStats['Fluid'] != 'DMSO') & (AllStats['Concentration'] == '20') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['cpaccuracy']) ] = 'FAIL'
            AllStats['Status'][(AllStats['Fluid'] != 'DMSO') & (AllStats['Concentration'] != '20') & (AllStats['RawCV']> crit['cvpct'])] = 'FAIL'
            AllStats['Status'][(AllStats['Fluid'] != 'DMSO') & (AllStats['Concentration'] != '20') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['accuracy']) ] = 'FAIL'
        else:
            AllStats['Status'][(AllStats['Fluid'] != 'DMSO') & (AllStats['RawCV']> crit['cvpct'])] = 'FAIL'
            AllStats['Status'][(AllStats['Fluid'] != 'DMSO') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['accuracy']) ] = 'FAIL'

        AllStats['Status'][AllStats['Reliability'] > crit['reliabilities']]  = 'FAIL'
 #       AllStats['Status'][AllStats['Empties'] > empties]  = 'FAIL'
        AllStats['Precision'] = 1
        AllStats['Accuracy'] = 1

        AllStats['Precision'][(AllStats['Fluid'] == 'DMSO') & (AllStats['RawCV']>crit['dmsocvpct'] )] = 0
        AllStats['Accuracy'][(AllStats['Fluid'] == 'DMSO') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['dmsoaccuracy']) ] = 0
        AllStats['Precision'][(AllStats['Fluid'] != 'DMSO') & (AllStats['RawCV']> crit['cvpct'] )] = 0
        AllStats['Accuracy'][(AllStats['Fluid'] != 'DMSO') & ((abs(AllStats['Mean']-nomvol) / nomvol )> crit['accuracy']) ] = 0         

#        print(AllStats)
        fbase = self.OutPutFolder.get() + './' + self.EchoSN.get() + '_' + prefix + '_' + self.passfailCriteria.get()
        try:
             processCCT.writeReport(AllStats,xlfilebase=fbase, crit=crit)
        except:
             messagebox.showerror('Error', ' Problem writing report ' + fbase)
             return -1
        return 0
    
    def goodReaderCombo(self, attributes):
        supportedLUTS = self.ReaderDat.getSupportedLUTs(self.selectedReader.get(),UseInternalStd=self.useInternalStd.get())
        if attributes['LUT'].get() not in supportedLUTS:
            messagebox.showerror("Error", "Error parsing File:: The LUTDATE " + attributes['LUT'].get() + " Not supported for reader " + self.selectedReader.get() +" if useInternalStd = " + str(self.useInternalStd.get()))
            return False
        return True
        
                   
        
    def determinePlateType(self, plastic = None, fluid = None, conc = None):
        platetype = '384PPL_AQ_BP2'
        possibleplates = [ p for p in self.processES.plates if plastic in p ]
        if len(possibleplates) == 0:
            return platetype
        if 'DMSO' in fluid:
            sel = [ r for r in possibleplates if 'DMSO' in r ]
        elif 'SP' in fluid:
            sel = [ r for r in possibleplates if 'SP' in r ]
        elif 'BP/GP' in fluid:
            try:
                c = int(conc)
            except:
                c = 0
            if (c > 30 ) or ('96' in plastic) :
                sel = [ r for r in possibleplates if 'GP' in r]
            else:    
                sel = [ r for r in possibleplates if 'BP' in r]
        elif 'CP' in fluid:
            sel = [ r for r in possibleplates if 'CP' in r]
        elif 'AQ' in fluid:
            if 'LDV' in plastic:
                sel =  [ r for r in possibleplates if 'AQ' in r]
            elif '384PP' in plastic:
                sel = [ r for r in possibleplates if 'CP' in r ]
            else:
                sel = [ r for r in possibleplates if 'AQ' in r]
        print('possible plates: ' , sel)
        if len(sel) > 0 :
            return sel[0]        
                
        return platetype
        
    def btnQuitClick(self):
#         for k, v in self.ReaderFileAttributes.items():
#            print(k , ' ::  ')
#            for key, value in v.items():
#                print('  ', key, ' : ', value.get())
    
         root.destroy()
        

root = tkinter.Tk()
root.title("pyDropCV v4.0")
#http://infohost.nmt.edu/tcc/help/pubs/tkinter/std-attrs.html#geometry
#http://infohost.nmt.edu/tcc/help/pubs/tkinter/toplevel.html
root.geometry("1300x800+10+10")

gui = TheGui(root)
root.mainloop()